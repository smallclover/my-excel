import calendar
import os

import openpyxl
import jpholiday
from datetime import datetime, timedelta
import win32com.client as win32
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from db_op import DBOperations
from small_tools import SmallTools


class ExcelUtil:

    def export_to_excel(self, export_date, input_file='template_3.xlsm', output_file='./wangshun.xlsx',
                        sheet_name='勤務報告'):

        print(f'export_date: {export_date}')
        # 读取现有工作簿
        wb = openpyxl.load_workbook(input_file)
        ws = wb.active

        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            raise ValueError(f"Sheet '{sheet_name}' does not exist in the workbook.")

        # 获取当前日期
        # today = datetime.today()
        # first_day = today.replace(day=1)
        # next_month = first_day.replace(day=28) + timedelta(days=4)
        # last_day = next_month - timedelta(days=next_month.day)

        # 将 export_date 转换为 datetime 对象
        export_year, export_month = map(int, export_date.split('/'))

        # 获取该月的第一天
        export_first_day = datetime(export_year, export_month, 1)

        # 获取该月的最后一天
        # 获取下个月的第一天，再减去一天得到本月最后一天
        if export_month == 12:
            export_last_day = datetime(export_year + 1, 1, 1) - timedelta(days=1)
        else:
            export_last_day = datetime(export_year, export_month + 1, 1) - timedelta(days=1)

        # 获取开始单元格的行和列
        start_row, start_col = openpyxl.utils.cell.coordinate_to_tuple('B10')

        # 获取源单元格的格式
        source_date_cell = ws.cell(row=start_row, column=start_col)
        source_weekday_cell = ws.cell(row=start_row, column=start_col + 1)
        source_relax_cell = ws.cell(row=start_row, column=start_col + 2)

        # 定义日本曜日列表
        japanese_weekdays = ['月', '火', '水', '木', '金', '土', '日']

        def get_relax_cell_value(date, name):
            """根据日期和星期名称获取D列单元格的值"""
            if jpholiday.is_holiday(date):
                return '祝日'  # 节假日
            elif name in ['土', '日']:
                return '休日'  # 周六或周日
            else:
                return ''  # 空白

        def copy_cell_format(source_cell, target_cell):
            target_cell.font = source_cell.font.copy()
            target_cell.fill = source_cell.fill.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.alignment = source_cell.alignment.copy()

        current_date = export_first_day
        while current_date <= export_last_day:
            date_cell = ws.cell(row=start_row, column=start_col, value=current_date)
            copy_cell_format(source_date_cell, date_cell)
            date_cell.number_format = source_date_cell.number_format

            # 填写对应的日本曜日
            weekday_cell_val = japanese_weekdays[current_date.weekday()]
            weekday_cell = ws.cell(row=start_row, column=start_col + 1, value=weekday_cell_val)
            copy_cell_format(source_weekday_cell, weekday_cell)

            # 获取D列的值
            relax_cell_val = get_relax_cell_value(current_date, weekday_cell_val)
            relax_cell = ws.cell(row=start_row, column=start_col + 2, value=relax_cell_val)
            copy_cell_format(source_relax_cell, relax_cell)

            current_date += timedelta(days=1)
            start_row += 1  # 向下移动到下一行

        # 获取合并单元格的起始单元格
        e2_cell = ws['E2']

        # 将值设置为 datetime 对象
        # 如果是设置字符串他并不会对格式生效
        # 值的类型会影响到格式，请注意
        e2_cell.value = datetime.strptime(export_first_day.strftime('%Y/%m/%d'), '%Y/%m/%d')

        # 恢复样式
        e2_cell.font = e2_cell.font.copy()
        e2_cell.fill = e2_cell.fill.copy()
        e2_cell.border = e2_cell.border.copy()
        e2_cell.alignment = e2_cell.alignment.copy()
        e2_cell.number_format = '[$-411]ggge"年"m"月"'
        print(e2_cell.number_format)

        data_list = DBOperations.get_data(export_date)

        # 解析 export_date
        # export_date = "2024/08"
        year, month = map(int, export_date.split('/'))

        # 获取月份的天数
        days_in_month = calendar.monthrange(year, month)[1]

        # 遍历每一天
        for day in range(1, days_in_month + 1):
            current_date = datetime(year, month, day)

            # 检查 data_list 中是否有当前日期的数据
            date_exists_in_data_list = any(
                datetime.strptime(row[1], "%Y/%m/%d").date() == current_date.date() for row in data_list)

            if SmallTools.is_holiday_or_weekend(current_date) and not date_exists_in_data_list:
                continue  # 如果是休日或祝日，并且 data_list 中没有这个日期的数据，则跳过这一天

            # 查找data_list中是否有当前日期的数据
            for row in data_list:
                data_date = datetime.strptime(row[1], "%Y/%m/%d")  # 假设 row[1] 是日期
                # print(f'data_date: {data_date}, current_date: {current_date}')
                if data_date.date() == current_date.date():
                    try:
                        start_time = datetime.strptime(row[2], "%Y/%m/%d %H:%M:%S")
                        end_time = datetime.strptime(row[3], "%Y/%m/%d %H:%M:%S")
                        time_diff = end_time - start_time
                        time_diff_hours = time_diff.seconds // 3600  # 转换为小时并取整
                        rest_hours = int(row[4]) if row[4] else 0
                        i_value = str(time_diff_hours - rest_hours)
                    except ValueError:
                        i_value = ""

                    # 填充Excel表格
                    idx = current_date.day - 1  # 根据日期计算行索引
                    ws.cell(row=10 + idx, column=openpyxl.utils.cell.column_index_from_string('E'),
                            value=start_time)  # 开始时间
                    ws.cell(row=10 + idx, column=openpyxl.utils.cell.column_index_from_string('F'),
                            value=end_time)  # 结束时间
                    ws.cell(row=10 + idx, column=openpyxl.utils.cell.column_index_from_string('G'),
                            value=row[4])  # 休息时间
                    ws.cell(row=10 + idx, column=openpyxl.utils.cell.column_index_from_string('I'),
                            value=i_value)  # 时间差
                    ws.cell(row=10 + idx, column=openpyxl.utils.cell.column_index_from_string('J'), value=row[5])  # 内容

                    break  # 只处理当前日期的数据

        # 插入图片到L45单元格        # img_path = 'modified_stamp_1.png'  # 这是你之前生成并保存的图片路径
        #         # img = ExcelImage(img_path)
        #         #
        #         # # 设置图片位置为L45单元格
        #         # img.anchor = 'L45'
        #         #
        #         # # 将图片添加到工作表
        #         # ws.add_image(img)

        # 读取 J 列第 1 到第 10 行的数据
        column = 'I'
        start_row = 10
        end_row = 40

        total = 0
        for row in range(start_row, end_row + 1):
            cell_value = ws[f"{column}{row}"].value
            if cell_value is not None:
                total += int(cell_value)

        # 将总和写入 J11
        ws[f"{column}41"] = total
        # 保存到新的文件
        wb.save(output_file)

        self.excel_to_pdf('wangshun.xlsx', '勤務報告', 'output.pdf')

    def excel_to_pdf(self, excel_file, sheet_name, pdf_file):
        # 获取当前工作目录
        current_directory = os.path.dirname(os.path.abspath(__file__))

        # 组合文件的完整路径
        excel_path = os.path.join(current_directory, excel_file)
        pdf_path = os.path.join(current_directory, pdf_file)

        # 创建Excel应用程序对象
        excel = win32.Dispatch('Excel.Application')

        # 打开Excel文件
        workbook = excel.Workbooks.Open(excel_path)

        # 获取指定的工作表
        sheet = workbook.Sheets(sheet_name)

        # 禁用Excel的警告弹窗
        excel.DisplayAlerts = False

        # 将工作表导出为PDF
        sheet.ExportAsFixedFormat(0, pdf_path)

        # 关闭工作簿
        workbook.Close(SaveChanges=False)

        # 退出Excel应用程序
        excel.Quit()
